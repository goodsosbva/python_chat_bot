import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active;
sheet.title = "회원정보"

header_title = ['이름', '전화번호']
for idx, title in enumerate(header_title):
    sheet.cell(row=1, column=idx+1, value=title)

member_info = [
    ('kei', '010-1234-5678'),
    ('hong', '010-1234-5679'),
]
row_num = 2
for row, member in enumerate(member_info):
    for col_num, v in enumerate(member):
        sheet.cell(row=row_num + row, column=col_num+1, value=v)

wb.save('kh_member.xlsx')
wb.close()


